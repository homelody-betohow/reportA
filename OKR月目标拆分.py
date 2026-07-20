"""
月度 OKR 目标拆解工具

流程：
  1. 从年度源表按月份、必填列筛选并复制指定 sheet
  2. 生成当月文件，如 2026-06月目标拆解及跟进.xlsx
  3. 合并 ALL：仅 1 个表头，按表头名对齐列，可带来源 sheet 列

用法：
  配置 OKR_MONTH = "yyyy-mm" 后运行: python monthOKR.py
  python monthOKR.py --merge-only
"""

import argparse
import openpyxl
from openpyxl import Workbook, load_workbook
import os
import re
import sys

# ========== 配置（按需修改） ==========
EXCEL_DIR = r"F:\月目标拆解及跟进"
SOURCE_FILENAME = "2026年目标拆解.xlsx"
OKR_MONTH = "2026-07"

SHEETS_TO_COPY = [
    "AMAZON-EU",
    "AMAZON-US",
    "REAL",
    "OTTO",
    "DLZ",
    "LM-TOTO",
    "LM-BTH",
    "TEMU-BV",
    "TEMU-BZ",
    "TEMU-AIH",
    "TEMU-HM",
    "MANO-OHPA",
    "MANO-COM",
]

MONTH_COLUMN_NAMES = ["月份", "日期", "时间", "Month", "Date"]
ENABLE_MONTH_FILTER = True

NOT_EMPTY_COLUMN_GROUPS = [
    ["识别SKU"],
    ["识别商品ID"],
]
ENABLE_NOT_EMPTY_FILTER = True
# 必填列在表头中必须全部存在，否则报错（避免静默跳过筛选）
REQUIRE_ALL_NOT_EMPTY_COLUMNS = True

MERGE_SHEET_NAME = "ALL"
MERGE_ALIGN_BY_HEADER = True       # 按表头名对齐列，避免各 sheet 列顺序不同导致错位
MERGE_ADD_SOURCE_COLUMN = True     # ALL 表增加「来源Sheet」列
SOURCE_COLUMN_NAME = "来源Sheet"
SKIP_DUPLICATE_HEADER_ROWS = True  # 跳过与表头相同的重复行
SKIP_EMPTY_ROWS = True             # 跳过整行为空的行


# ---------- 路径与月份 ----------


def get_source_path(excel_dir=None):
    return os.path.join(excel_dir or EXCEL_DIR, SOURCE_FILENAME)


def get_okr_month():
    env = os.environ.get("OKR_MONTH", "").strip()
    return env if env else (OKR_MONTH or "").strip()


def parse_month(month_str):
    s = month_str.strip().replace("月", "").replace("年", "-").replace(".", "-")
    s = re.sub(r"\s+", "", s)
    if re.fullmatch(r"\d{2}-\d{1,2}", s):
        y, m = s.split("-")
        year, month = 2000 + int(y), int(m)
    elif re.fullmatch(r"\d{4}-\d{1,2}", s):
        y, m = s.split("-")
        year, month = int(y), int(m)
    else:
        raise ValueError(f"无法识别的月份格式: {month_str}")
    if not 1 <= month <= 12:
        raise ValueError(f"月份须在 1-12 之间: {month}")
    return year, month


def get_monthly_filename(year, month):
    return f"{year}-{month:02d}月目标拆解及跟进.xlsx"


def get_monthly_path(year, month, excel_dir=None):
    return os.path.join(excel_dir or EXCEL_DIR, get_monthly_filename(year, month))


# ---------- 文件与表头工具 ----------


def is_file_writable(path):
    if not os.path.isfile(path):
        return True
    try:
        with open(path, "r+b"):
            pass
        return True
    except (PermissionError, OSError):
        return False


def print_file_locked_error(excel_path):
    print(f"错误：文件被占用，无法写入 - {excel_path}")
    print("请关闭 Excel/WPS 中打开的该文件后重试。")


def normalize_header(cell):
    if cell is None:
        return ""
    return str(cell).strip()


def find_column(header_row, column_names):
    for name in column_names:
        for idx, cell_value in enumerate(header_row):
            if cell_value is None:
                continue
            if str(cell_value).strip() == name:
                return idx
    for name in column_names:
        for idx, cell_value in enumerate(header_row):
            if cell_value is None:
                continue
            if name in str(cell_value).strip():
                return idx
    return None


def resolve_not_empty_columns(header_row, column_groups=None):
    column_groups = column_groups or NOT_EMPTY_COLUMN_GROUPS
    resolved = []
    for names in column_groups:
        idx = find_column(header_row, names)
        if idx is not None:
            resolved.append((idx, header_row[idx]))
        else:
            resolved.append((None, names[0]))
    return resolved


def is_cell_not_empty(cell_value):
    if cell_value is None:
        return False
    if isinstance(cell_value, str):
        return bool(cell_value.strip())
    return True


def is_row_empty(row):
    return not any(is_cell_not_empty(v) for v in row)


def is_duplicate_header_row(row, canonical_header):
    if not SKIP_DUPLICATE_HEADER_ROWS or not canonical_header:
        return False
    for i, h in enumerate(canonical_header):
        rv = row[i] if i < len(row) else None
        if normalize_header(rv) != normalize_header(h):
            return False
    return True


def align_row_to_header(row, source_header, canonical_header):
    """按 canonical_header 列名从 source 行取值。"""
    src_index = {normalize_header(h): i for i, h in enumerate(source_header) if normalize_header(h)}
    out = []
    for h in canonical_header:
        key = normalize_header(h)
        if key in src_index and src_index[key] < len(row):
            out.append(row[src_index[key]])
        else:
            out.append(None)
    return out


def match_month_value(cell_value, target_year, target_month):
    if cell_value is None:
        return False
    if hasattr(cell_value, "year") and hasattr(cell_value, "month"):
        return cell_value.year == target_year and cell_value.month == target_month
    s = str(cell_value).strip().replace("月", "").replace("年", "-").replace("/", "-")
    m = re.match(r"(\d{4})-(\d{1,2})", s)
    if m:
        y, mon = int(m.group(1)), int(m.group(2))
        return y == target_year and mon == target_month
    m = re.match(r"(\d{4})(\d{2})", s)
    if m:
        y, mon = int(m.group(1)), int(m.group(2))
        return y == target_year and mon == target_month
    return False


# ---------- 筛选上下文（避免每行重复查找列） ----------


def build_filter_context(header_row, year, month):
    use_month = ENABLE_MONTH_FILTER and year is not None and month is not None
    use_not_empty = ENABLE_NOT_EMPTY_FILTER

    month_col_idx = find_column(header_row, MONTH_COLUMN_NAMES) if use_month else None
    if use_month and month_col_idx is None:
        use_month = False

    not_empty_cols = resolve_not_empty_columns(header_row) if use_not_empty else []
    missing = [label for idx, label in not_empty_cols if idx is None]
    active_not_empty = [(idx, label) for idx, label in not_empty_cols if idx is not None]

    if use_not_empty and missing:
        msg = f"未找到必填列: {', '.join(missing)}"
        if REQUIRE_ALL_NOT_EMPTY_COLUMNS:
            raise ValueError(msg)
        print(f"    警告：{msg}，跳过对应列校验")

    if use_not_empty and not active_not_empty and not missing:
        use_not_empty = False

    return {
        "header_row": header_row,
        "month_col_idx": month_col_idx,
        "active_not_empty": active_not_empty,
        "use_month": use_month,
        "use_not_empty": use_not_empty and bool(active_not_empty),
    }


def row_passes_filters(row, ctx, year, month):
    if ctx["use_month"]:
        idx = ctx["month_col_idx"]
        if idx >= len(row) or not match_month_value(row[idx], year, month):
            return False, "month"
    if ctx["use_not_empty"]:
        for col_idx, _ in ctx["active_not_empty"]:
            if col_idx >= len(row) or not is_cell_not_empty(row[col_idx]):
                return False, "not_empty"
    return True, None


def format_filter_log(ctx, header_row, copied, skip_month, skip_not_empty, skip_empty=0, skip_dup_header=0):
    parts = [f"复制 {copied} 行"]
    # if ctx["use_month"] and ctx["month_col_idx"] is not None:
    #     i = ctx["month_col_idx"]
    #     parts.append(f"月份列第 {i + 1} 列（{header_row[i]}）")
    #     if skip_month:
    #         parts.append(f"跳过月份不匹配 {skip_month} 行")
    # if ctx["use_not_empty"]:
    #     col_desc = "、".join(f"第 {i + 1} 列（{l}）" for i, l in ctx["active_not_empty"])
    #     parts.append(f"必填列 {col_desc}")
    #     if skip_not_empty:
    #         parts.append(f"跳过必填列为空 {skip_not_empty} 行")
    if skip_empty:
        parts.append(f"跳过空行 {skip_empty} 行")
    if skip_dup_header:
        parts.append(f"跳过重复表头 {skip_dup_header} 行")
    return "，".join(parts)


# ---------- 复制 sheet ----------


def copy_sheet_values(source_ws, target_ws, year=None, month=None):
    rows = list(source_ws.iter_rows(values_only=True))
    if not rows:
        return

    header_row = list(rows[0])
    target_ws.append(header_row)
    data_rows = rows[1:]

    if not (ENABLE_MONTH_FILTER or ENABLE_NOT_EMPTY_FILTER) or (year is None or month is None):
        for row in data_rows:
            if SKIP_EMPTY_ROWS and is_row_empty(row):
                continue
            target_ws.append(list(row))
        print(f"    未启用筛选，复制 {target_ws.max_row - 1} 行")
        return

    try:
        ctx = build_filter_context(header_row, year, month)
    except ValueError as e:
        print(f"    错误：{e}")
        return

    copied = skip_month = skip_not_empty = skip_empty = 0
    for row in data_rows:
        if SKIP_EMPTY_ROWS and is_row_empty(row):
            skip_empty += 1
            continue
        ok, reason = row_passes_filters(row, ctx, year, month)
        if not ok:
            if reason == "month":
                skip_month += 1
            else:
                skip_not_empty += 1
            continue
        target_ws.append(list(row))
        copied += 1

    print(f"    {format_filter_log(ctx, header_row, copied, skip_month, skip_not_empty, skip_empty)}")


def build_monthly_workbook(year, month, sheets_to_copy=None, excel_dir=None):
    sheets_to_copy = sheets_to_copy or SHEETS_TO_COPY
    source_path = get_source_path(excel_dir)
    output_path = get_monthly_path(year, month, excel_dir)

    if not os.path.exists(source_path):
        print(f"错误：年度源文件不存在 - {source_path}")
        return None
    if not is_file_writable(output_path):
        print_file_locked_error(output_path)
        return None
    if not sheets_to_copy:
        print("错误：SHEETS_TO_COPY 为空")
        return None

    src_wb = out_wb = None
    try:
        print(f"正在读取年度源表: {source_path}")
        src_wb = load_workbook(source_path, read_only=True, data_only=True)
        out_wb = Workbook()
        out_wb.remove(out_wb.active)

        copied, missing = [], []
        for sheet_name in sheets_to_copy:
            if sheet_name not in src_wb.sheetnames:
                missing.append(sheet_name)
                continue
            print(f"  正在复制 sheet: {sheet_name}")
            target_ws = out_wb.create_sheet(sheet_name)
            copy_sheet_values(src_wb[sheet_name], target_ws, year, month)
            copied.append(sheet_name)

        if missing:
            print(f"  警告：源表缺少 sheet: {', '.join(missing)}")
            print(f"        现有: {', '.join(src_wb.sheetnames)}")
        if not copied:
            print("错误：未复制任何 sheet")
            return None

        print(f"\n正在生成: {output_path}")
        out_wb.save(output_path)
        print(f"✓ 已生成 {get_monthly_filename(year, month)}，{len(copied)} 个 sheet")
        return output_path
    except Exception as e:
        print(f"错误：{e}")
        import traceback
        traceback.print_exc()
        return None
    finally:
        if src_wb:
            src_wb.close()
        if out_wb:
            out_wb.close()


# ---------- 合并 ALL（仅 1 个表头） ----------


def merge_sheets_to_all(excel_path, merge_sheet_name=None, year=None, month=None):
    merge_sheet_name = merge_sheet_name or MERGE_SHEET_NAME

    if not os.path.exists(excel_path):
        print(f"错误：文件不存在 - {excel_path}")
        return False
    if not is_file_writable(excel_path):
        print_file_locked_error(excel_path)
        return False

    wb = None
    try:
        print(f"\n正在合并 ALL: {excel_path}")
        wb = load_workbook(excel_path)
        all_sheets = [s for s in wb.sheetnames if s != merge_sheet_name]
        if not all_sheets:
            print("错误：没有可合并的 sheet")
            return False

        print(f"待合并: {', '.join(all_sheets)}")
        if merge_sheet_name in wb.sheetnames:
            del wb[merge_sheet_name]

        all_sheet = wb.create_sheet(merge_sheet_name, 0)
        master_header = None  # 第一个 sheet 的表头，作为 ALL 列标准
        data_row_count = 0
        total_skip = {"month": 0, "not_empty": 0, "empty": 0, "dup_header": 0}

        for sheet_name in all_sheets:
            sheet = wb[sheet_name]
            max_row, max_col = sheet.max_row or 0, sheet.max_column or 0
            if max_row < 1:
                print(f"  - {sheet_name}: 空，跳过")
                continue

            header_row = list(
                sheet.iter_rows(min_row=1, max_row=1, max_col=max_col, values_only=True)
            )[0]

            if master_header is None:
                master_header = list(header_row)
                all_header = master_header + (
                    [SOURCE_COLUMN_NAME] if MERGE_ADD_SOURCE_COLUMN else []
                )
                for col_idx, h in enumerate(all_header, start=1):
                    all_sheet.cell(1, col_idx, h)
                print(
                    f"  ALL 表头（{len(all_header)} 列，仅写入 1 次）: "
                    f"{', '.join(str(h) for h in all_header[:8])}..."
                )

            try:
                ctx = build_filter_context(header_row, year, month)
            except ValueError as e:
                print(f"  - {sheet_name}: 跳过，{e}")
                continue

            rows_copied = 0
            skip = {"month": 0, "not_empty": 0, "empty": 0, "dup_header": 0}
            base_header = list(header_row)

            for row_values in sheet.iter_rows(
                min_row=2, max_row=max_row, max_col=max_col, values_only=True
            ):
                row = list(row_values)

                if SKIP_EMPTY_ROWS and is_row_empty(row):
                    skip["empty"] += 1
                    continue
                if is_duplicate_header_row(row, base_header):
                    skip["dup_header"] += 1
                    continue

                ok, reason = row_passes_filters(row, ctx, year, month)
                if not ok:
                    skip[reason] += 1
                    continue

                if MERGE_ALIGN_BY_HEADER and master_header:
                    out_row = align_row_to_header(row, base_header, master_header)
                else:
                    out_row = list(row)
                    if len(out_row) < len(master_header):
                        out_row.extend([None] * (len(master_header) - len(out_row)))
                    elif len(out_row) > len(master_header):
                        out_row = out_row[: len(master_header)]
                if MERGE_ADD_SOURCE_COLUMN:
                    out_row = out_row + [sheet_name]

                data_row_count += 1
                rows_copied += 1
                all_sheet.append(out_row)

            msg = f"  - {sheet_name}: {rows_copied} 行"
            if any(skip.values()):
                extras = []
                if skip["month"]:
                    extras.append(f"月份不匹配 {skip['month']}")
                if skip["not_empty"]:
                    extras.append(f"必填列为空 {skip['not_empty']}")
                if skip["empty"]:
                    extras.append(f"空行 {skip['empty']}")
                if skip["dup_header"]:
                    extras.append(f"重复表头 {skip['dup_header']}")
                msg += f"（跳过: {', '.join(extras)}）"
            print(msg)
            for k in total_skip:
                total_skip[k] += skip[k]

        if master_header is None:
            print("错误：未写入任何数据")
            return False

        for col_idx in range(1, (all_sheet.max_column or 0) + 1):
            letter = openpyxl.utils.get_column_letter(col_idx)
            all_sheet.column_dimensions[letter].width = 15

        wb.save(excel_path)
        print(
            f"✓ ALL 完成：1 行表头 + {data_row_count} 行数据"
            f"（共 {all_sheet.max_row} 行）"
        )
        if any(total_skip.values()):
            print(f"  筛选跳过合计: {total_skip}")
        return True

    except (PermissionError, OSError) as e:
        if isinstance(e, PermissionError) or getattr(e, "errno", None) == 13:
            print_file_locked_error(excel_path)
            return False
        raise
    except Exception as e:
        print(f"错误：合并失败 - {e}")
        import traceback
        traceback.print_exc()
        return False
    finally:
        if wb:
            wb.close()


# ---------- 入口 ----------


def prompt_month():
    while True:
        raw = input("请输入月份（如 2026-06）: ").strip()
        if not raw:
            print("月份不能为空")
            continue
        try:
            return parse_month(raw)
        except ValueError as e:
            print(f"{e}")


def run_pipeline(year, month, merge_only=False, excel_dir=None):
    print("=" * 60)
    print("月度 OKR 目标拆解工具")
    print("=" * 60)
    print(f"月份: {year}-{month:02d}月")
    print(f"输出: {get_monthly_filename(year, month)}")
    print()

    excel_path = get_monthly_path(year, month, excel_dir)
    if not merge_only:
        excel_path = build_monthly_workbook(year, month, excel_dir=excel_dir)
        if not excel_path:
            return False
    elif not os.path.exists(excel_path):
        print(f"错误：文件不存在 - {excel_path}")
        return False

    return merge_sheets_to_all(excel_path, year=year, month=month)


def main():
    parser = argparse.ArgumentParser(description="从年度表生成当月 OKR 并合并 ALL")
    parser.add_argument("--merge-only", action="store_true", help="仅合并 ALL")
    parser.add_argument("--dir", default=None, help=f"工作目录，默认 {EXCEL_DIR}")
    args = parser.parse_args()
    excel_dir = args.dir or EXCEL_DIR

    try:
        month_str = get_okr_month()
        year, month = parse_month(month_str) if month_str else prompt_month()
    except ValueError as e:
        print(f"错误：{e}")
        sys.exit(1)

    if run_pipeline(year, month, merge_only=args.merge_only, excel_dir=excel_dir):
        print("\n✓ 全部处理完成！")
        print("=" * 60)
    else:
        print("\n✗ 处理失败")
        sys.exit(1)


if __name__ == "__main__":
    main()
