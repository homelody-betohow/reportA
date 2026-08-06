"""
K0_库存周转.py — 从 snapshot_market_turnover 反向生成库存动销明细 Excel

输出：
  {DESKTOP_ROOT}\\{folder_name}{shared_date}\\仓租\\{yyyy.m.d}库存动销明细.xlsx
  sheet：
    1. 各平台SKU库存动销明细
    2. 各平台商品库存周转明细（由 sheet1 按 平台+商品ID 汇总，单价货值除外；不含销售站点）
    3. 商品库存周转明细（由 sheet1 按 商品ID 汇总，单价货值除外；不含销售平台、销售站点）
    4. 库存周转汇总透视（三块：按平台 / 产品状态 / 销售负责人；含货值×数量）

供 K1/K2 仓租分摊读取（需含列：商品ID、销售平台、SKU、可售库存-可调）。

运营负责人（AMAZON）：口径对齐 M3_映射_销售负责人_AMZ —
  AMAZON-EU：月目标表按商品ID→负责人，再用「信息-映射」SKU 覆盖；空→nobody
  AMAZON-US：月目标表按商品ID→负责人；SKU 以 U 开头→官雪婷US；空→nobody
  其余平台：保留 snapshot_market_turnover.ops_owner（「无负责人」/空 → nobody）

用法：
  python modules/K_storage_fee_product_info/K0_库存周转.py
  python modules/K_storage_fee_product_info/K0_库存周转.py --date 2026.7.18
  python modules/K_storage_fee_product_info/K0_库存周转.py -d 2026-07-18
"""

from __future__ import annotations

import argparse
import importlib.util
import re
import warnings
from collections.abc import Callable
from datetime import date, datetime
from pathlib import Path

import pandas as pd
import pymysql.cursors
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

_epr_file = next(
    p / "ensure_project_root.py"
    for p in Path(__file__).resolve().parents
    if (p / "ensure_project_root.py").is_file()
)
_spec = importlib.util.spec_from_file_location("ensure_project_root", _epr_file)
_epr_mod = importlib.util.module_from_spec(_spec)
_spec.loader.exec_module(_epr_mod)
_epr_mod.bootstrap(__file__)

from common.sku_mapping import sku_mappings  # noqa: E402
from config.A0_paths import DESKTOP_ROOT, MONTH_GOAL_EXCEL_PATH  # noqa: E402
from config.A0_set_date import folder_name, ku_cun_date, report_date, shared_date  # noqa: E402
from database.db_connection import get_db_manager  # noqa: E402

# ---------------------------------------------------------------------------
# 常量 / SQL / 样式
# ---------------------------------------------------------------------------

_INFO_MAP_XLSX = fr"{DESKTOP_ROOT}\信息-映射.xlsx"

warnings.filterwarnings("ignore", category=UserWarning, module="openpyxl.styles.stylesheet")

SHEET_SKU = "各平台SKU库存周转明细"
SHEET_PRODUCT_BY_MARKET = "各平台商品ID周转明细"
SHEET_PRODUCT = "商品ID周转明细"
SHEET_PIVOT = "库存周转汇总"
FILE_SUFFIX = "库存动销明细.xlsx"

_NA_TOKENS = frozenset({"", "nan", "None", "NaN"})
_STATUS_BLANK = frozenset({"", "无", "--", "nan", "None", "NaN"})
_STATUS_FILL = "无"
_OWNER_BLANK = frozenset({"", "nan", "None", "NaN", "无负责人"})
_OWNER_FILL = "nobody"

# 商品级汇总时不求和的列（单价货值取首行；周转/货值衍生列汇总后重算）
_PRODUCT_SKIP_SUM = frozenset(
    {
        "货值/个",
        "海外周转-月",
        "总库存周转-月",
        "销售货值",
        "海外库存货值",
        "海外货值周转",
    }
)
# 商品级汇总时取首行的文本/标识列（非分组键时生效）
_PRODUCT_FIRST_COLS = (
    "销售平台",
    "商品ID",
    "SKU",
    "产品状态",
    "运营经理",
    "运营负责人",
    "货值/个",
    "销售站点",
    "供应商",
)

_EXPORT_COLS = (
    "销售平台",
    "销售站点",
    "商品ID",
    "SKU",
    "供应商",
    "产品状态",
    "运营经理",
    "运营负责人",
    "参考月销量",
    "计划库存-禁调",
    "在途库存-禁调",
    "可售库存-禁调",
    "计划库存-可调",
    "在途库存-可调",
    "可售库存-可调",
    "总计划库存",
    "总在途库存",
    "总可售库存",
    "总海外库存",
    "海外周转-月",
    "总库存",
    "总库存周转-月",
    "货值/个",
    "销售货值",
    "海外库存货值",
    "海外货值周转",
)

# 库存/销量：整数
_INT_COLS = (
    "计划库存-禁调",
    "在途库存-禁调",
    "可售库存-禁调",
    "计划库存-可调",
    "在途库存-可调",
    "可售库存-可调",
    "总计划库存",
    "总在途库存",
    "总可售库存",
    "参考月销量",
    "总海外库存",
    "总库存",
)
# 货值、周转天数：浮点
_FLOAT_COLS = (
    "货值/个",
    "海外周转-月",
    "总库存周转-月",
    "销售货值",
    "海外库存货值",
    "海外货值周转",
)
# 禁调列：灰白背景
_DIR_COLS = (
    "计划库存-禁调",
    "在途库存-禁调",
    "可售库存-禁调",
)
_DIR_FILL = PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid")
_HEADER_FILL = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
_PIVOT_HEADER_FILL = PatternFill(start_color="D9D9D9", end_color="D9D9D9", fill_type="solid")
_PIVOT_SALES_FILL = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")
_HEADER_FONT = Font(bold=True, color="FFFFFF", name="微软雅黑", size=10)
_PIVOT_HEADER_FONT = Font(bold=True, color="000000", name="微软雅黑", size=10)
_BODY_FONT = Font(name="微软雅黑", size=10)
_TOTAL_FONT = Font(bold=True, name="微软雅黑", size=10)
_HEADER_ALIGN = Alignment(horizontal="center", vertical="center", wrap_text=True)
_BODY_ALIGN = Alignment(horizontal="center", vertical="center")
_THIN_BORDER = Border(
    left=Side(style="thin", color="B4B4B4"),
    right=Side(style="thin", color="B4B4B4"),
    top=Side(style="thin", color="B4B4B4"),
    bottom=Side(style="thin", color="B4B4B4"),
)
_COL_WIDTH_MIN = 8
_COL_WIDTH_MAX = 28
_COL_WIDTH_PAD = 2

# 透视表：沿用 sheet1 原列名求和；另计货值金额（单价货值 × 数量）
_PIVOT_QTY_COLS = (
    "参考月销量",
    "计划库存-禁调",
    "在途库存-禁调",
    "可售库存-禁调",
    "计划库存-可调",
    "在途库存-可调",
    "可售库存-可调",
)
# 汇总后重算：海外=(D+E+G+H)/B；整体=(C+D+E+F+G+H)/B；B=0 → 0
_PIVOT_TURN_COLS = (
    "海外库存周转",
    "整体库存周转",
)
_PIVOT_AMOUNT_COLS = (
    "货值(在途+在库)",  # sum(货值/个 × 总海外库存)
    "货值(整体库存)",  # sum(货值/个 × 总库存)
)
_PIVOT_SUM_COLS = _PIVOT_QTY_COLS + _PIVOT_AMOUNT_COLS
_PIVOT_VALUE_COLS = _PIVOT_QTY_COLS + _PIVOT_AMOUNT_COLS + _PIVOT_TURN_COLS
_PIVOT_BLANK_LABELS = {
    "产品状态": _STATUS_FILL,
    "运营负责人": _OWNER_FILL,
}
_PIVOT_SECTIONS = (
    ("销售平台", "销售平台"),
    ("产品状态", "产品状态"),
    ("销售负责人", "运营负责人"),  # 表头用销售负责人，数据取运营负责人
)

# SQL 别名 → 导出列名
_RENAME_COLS = {
    "计划库存_禁调": "计划库存-禁调",
    "在途库存_禁调": "在途库存-禁调",
    "可售库存_禁调": "可售库存-禁调",
    "计划库存_可调拨": "计划库存-可调",
    "在途库存_可调": "在途库存-可调",
    "可售库存_可调": "可售库存-可调",
}

MARKET_TURNOVER_SQL = """
SELECT
    t.market_code          AS `销售市场`,
    t.product_uid          AS `商品ID`,
    t.product_sku          AS `SKU`,
    t.sku_lifecycle        AS `产品状态`,
    t.ops_leader           AS `运营经理`,
    t.ops_owner            AS `运营负责人`,
    t.supplier_name        AS `供应商`,
    t.cost_price_cny       AS `货值/个`,

    t.dir_planned_qty      AS `计划库存-禁调`,
    t.dir_onway_qty        AS `在途库存-禁调`,
    t.dir_sellable_qty     AS `可售库存-禁调`,

    t.trf_planned_qty      AS `计划库存-可调`,
    t.trf_onway_qty        AS `在途库存-可调`,
    t.trf_sellable_qty     AS `可售库存-可调`,

    t.total_planned_qty    AS `总计划库存`,
    t.total_onway_qty      AS `总在途库存`,
    t.total_sellable_qty   AS `总可售库存`,
    t.ref_month_sales_qty  AS `参考月销量`,
    t.market_region        AS `销售站点`,
    (t.total_onway_qty + t.total_sellable_qty) AS `总海外库存`,
    (t.total_onway_qty + t.total_sellable_qty) / t.ref_month_sales_qty AS `海外周转-月`,
    (t.total_onway_qty + t.total_sellable_qty + t.total_planned_qty) AS `总库存`,
    (t.total_onway_qty + t.total_sellable_qty + t.total_planned_qty) / t.ref_month_sales_qty AS `总库存周转-月`
FROM
    snapshot_market_turnover t
WHERE
    t.snapshot_date = %s
ORDER BY
    t.market_code, t.product_sku
"""


# ---------------------------------------------------------------------------
# 通用工具
# ---------------------------------------------------------------------------

def _is_blank_series(series: pd.Series, *, extra: frozenset[str] | set[str] = ()) -> pd.Series:
    """isna 或 strip 后属于空标记集合。"""
    tokens = _NA_TOKENS | frozenset(extra)
    return series.isna() | series.astype(str).str.strip().isin(tokens)


def _fill_blank_col(
    df: pd.DataFrame,
    col: str,
    fill,
    *,
    extra: frozenset[str] | set[str] = (),
) -> pd.DataFrame:
    """将 col 的空值替换为 fill；列不存在则原样返回。"""
    if col not in df.columns:
        return df
    out = df.copy()
    out.loc[_is_blank_series(out[col], extra=extra), col] = fill
    return out


def _num(
    df: pd.DataFrame | pd.Series,
    col: str | None = None,
    *,
    fill: float | int | None = 0,
) -> pd.Series:
    """to_numeric；fill 非 None 时 fillna。"""
    if isinstance(df, pd.Series):
        s = pd.to_numeric(df, errors="coerce")
    elif col is None or col not in df.columns:
        return pd.Series(fill if fill is not None else 0, index=getattr(df, "index", None))
    else:
        s = pd.to_numeric(df[col], errors="coerce")
    if fill is not None:
        return s.fillna(fill)
    return s


def _format_ku_cun_date(d: date) -> str:
    """与 A0_set_date.ku_cun_date 一致：yyyy.m.d（月/日不补零）。"""
    return f"{d.year}.{d.month}.{d.day}"


def parse_snapshot_date(raw: str) -> date:
    """
    解析快照日，支持：
      2026.7.18 / 2026.07.18
      2026-7-18 / 2026-07-18
      2026/7/18 / 20260718
    """
    s = str(raw).strip()
    if not s:
        raise ValueError("日期不能为空")

    if re.fullmatch(r"\d{8}", s):
        return datetime.strptime(s, "%Y%m%d").date()

    for fmt in ("%Y-%m-%d", "%Y/%m/%d", "%Y.%m.%d"):
        try:
            return datetime.strptime(s, fmt).date()
        except ValueError:
            pass

    # 兼容不补零：2026.7.18 / 2026-7-8
    m = re.fullmatch(r"(\d{4})[./-](\d{1,2})[./-](\d{1,2})", s)
    if m:
        y, mo, d = map(int, m.groups())
        return date(y, mo, d)

    raise ValueError(
        f"无法解析日期：{raw!r}，请使用如 2026.7.18 / 2026-07-18 / 20260718"
    )


def _display_width(value) -> int:
    """估算单元格显示宽度（中文约 2 个字符宽）。"""
    s = "" if value is None else str(value)
    width = 0
    for ch in s:
        width += 2 if ord(ch) > 127 else 1
    return width


def _put_platform_all_last(
    df: pd.DataFrame, *, col: str = "销售平台"
) -> pd.DataFrame:
    """销售平台 = ALL 的行放到末尾，其余行相对顺序不变。"""
    if df.empty or col not in df.columns:
        return df
    is_all = df[col].astype(str).str.strip().eq("ALL")
    if not is_all.any():
        return df
    return pd.concat([df.loc[~is_all], df.loc[is_all]], ignore_index=True)


# ---------------------------------------------------------------------------
# 日期与取数
# ---------------------------------------------------------------------------

def _fetch_market_turnover(snapshot_date: date) -> pd.DataFrame:
    db = get_db_manager()
    conn = db.get_connection()
    try:
        with conn.cursor(pymysql.cursors.DictCursor) as cur:
            cur.execute(MARKET_TURNOVER_SQL, (snapshot_date,))
            rows = cur.fetchall()
    finally:
        conn.close()

    if not rows:
        return pd.DataFrame(columns=list(_EXPORT_COLS))

    df = pd.DataFrame(rows)
    df = _attach_legacy_columns(df)
    df = _fill_blank_product_status(df)
    df = _map_amz_ops_owner(df)
    df = _coerce_numeric_columns(df)
    return _compute_value_metrics(df)


# ---------------------------------------------------------------------------
# 清洗与映射
# ---------------------------------------------------------------------------

def _attach_legacy_columns(df: pd.DataFrame) -> pd.DataFrame:
    """对齐导出列名，并补充 K1/K2 口径：销售平台。"""
    out = df.rename(columns=_RENAME_COLS).copy()
    if "销售平台" not in out.columns and "销售市场" in out.columns:
        out["销售平台"] = out["销售市场"]
    return out


def _fill_blank_product_id(df: pd.DataFrame) -> pd.DataFrame:
    """商品ID 为空时默认等于 SKU。"""
    if "商品ID" not in df.columns or "SKU" not in df.columns:
        return df
    out = df.copy()
    blank = _is_blank_series(out["商品ID"])
    out.loc[blank, "商品ID"] = out.loc[blank, "SKU"].astype(str)
    return out


def _fill_blank_product_status(df: pd.DataFrame) -> pd.DataFrame:
    """产品状态为空时赋值为 无。"""
    return _fill_blank_col(df, "产品状态", _STATUS_FILL, extra={"--"})


def _normalize_ops_owner(df: pd.DataFrame) -> pd.DataFrame:
    """库表/历史口径可能仍是「无负责人」或空，统一为 nobody。"""
    return _fill_blank_col(df, "运营负责人", _OWNER_FILL, extra={"无负责人"})


def _map_amz_eu_owner(eu: pd.DataFrame) -> pd.DataFrame:
    """AMAZON-EU：月目标商品ID→负责人，信息-映射 SKU 覆盖；空→nobody。"""
    eu = sku_mappings(
        main_df=eu,
        main_sku="商品ID",
        map_sku_path=MONTH_GOAL_EXCEL_PATH,
        map_old_sku="商品ID",
        map_new_sku="负责人",
        map_sku_sheet="AMAZON-EU",
    )
    eu = sku_mappings(
        main_df=eu,
        main_sku="SKU",
        map_sku_path=_INFO_MAP_XLSX,
        map_old_sku="SKU",
        map_new_sku="销售负责人-SKU（AMAZON-EU）",
        map_sku_sheet="销售负责人",
    )
    override = eu["映射销售负责人-SKU（AMAZON-EU）"].notna()
    eu.loc[override, "映射负责人"] = eu.loc[override, "映射销售负责人-SKU（AMAZON-EU）"]
    eu["运营负责人"] = eu["映射负责人"].fillna(_OWNER_FILL)
    return eu.drop(
        columns=["映射负责人", "映射销售负责人-SKU（AMAZON-EU）"],
        errors="ignore",
    )


def _map_amz_us_owner(us: pd.DataFrame) -> pd.DataFrame:
    """AMAZON-US：月目标商品ID→负责人；SKU 以 U 开头→官雪婷US；空→nobody。"""
    us = sku_mappings(
        main_df=us,
        main_sku="商品ID",
        map_sku_path=MONTH_GOAL_EXCEL_PATH,
        map_old_sku="商品ID",
        map_new_sku="负责人",
        map_sku_sheet="AMAZON-US",
    )
    us["运营负责人"] = us["映射负责人"]
    us.loc[us["SKU"].astype(str).str.startswith("U", na=False), "运营负责人"] = "官雪婷US"
    us["运营负责人"] = us["运营负责人"].fillna(_OWNER_FILL)
    return us.drop(columns=["映射负责人"], errors="ignore")


def _map_amz_ops_owner(df: pd.DataFrame) -> pd.DataFrame:
    """
    重写 AMAZON-EU / AMAZON-US 的「运营负责人」，逻辑对齐 M3_映射_销售负责人_AMZ。
    非 AMZ 行保留库表 ops_owner；中间映射列不落盘。
    """
    if df.empty or "销售平台" not in df.columns:
        return df
    if "商品ID" not in df.columns or "SKU" not in df.columns:
        return df

    platform = df["销售平台"]
    eu_mask = platform.isin(["AMAZON-EU"])
    us_mask = platform.isin(["AMAZON-US"])
    other_mask = ~(eu_mask | us_mask)
    parts: list[pd.DataFrame] = []

    if eu_mask.any():
        parts.append(_map_amz_eu_owner(df.loc[eu_mask].copy()))
    if us_mask.any():
        parts.append(_map_amz_us_owner(df.loc[us_mask].copy()))
    if other_mask.any():
        parts.append(df.loc[other_mask].copy())

    if not parts:
        return df
    # 按原索引还原 SQL 排序，再统一 reset
    out = pd.concat(parts).sort_index().reset_index(drop=True)
    return _normalize_ops_owner(out)


def _compute_value_metrics(df: pd.DataFrame) -> pd.DataFrame:
    """
    销售货值 = 货值/个 × 参考月销量
    海外库存货值 = 货值/个 × 总海外库存
    海外货值周转 = 海外库存货值 / 销售货值（分母为 0 → NaN）
    """
    out = df.copy()
    unit = _num(out, "货值/个")
    sales = _num(out, "参考月销量")
    overseas = _num(out, "总海外库存")
    out["销售货值"] = unit * sales
    out["海外库存货值"] = unit * overseas
    ref_safe = out["销售货值"].mask(out["销售货值"].eq(0))
    out["海外货值周转"] = out["海外库存货值"] / ref_safe
    return out


def _coerce_numeric_columns(df: pd.DataFrame) -> pd.DataFrame:
    """确保库存/天数/货值以数字类型写入 Excel（避免 Decimal/object 被当成文本）。"""
    out = df.copy()
    for col in _INT_COLS:
        if col not in out.columns:
            continue
        out[col] = _num(out, col).astype("int64")
    for col in _FLOAT_COLS:
        if col not in out.columns:
            continue
        out[col] = _num(out, col, fill=None).astype("float64")
    return out


# ---------------------------------------------------------------------------
# 商品汇总
# ---------------------------------------------------------------------------

def _first_prefer_status(series: pd.Series) -> str:
    """分组取产品状态：优先非 无，全空则 无。"""
    vals = series.astype(str).str.strip()
    non_blank = vals[~vals.isin(_STATUS_BLANK)]
    if not non_blank.empty:
        return str(non_blank.iloc[0])
    return _STATUS_FILL


def _recompute_inventory_metrics(grouped: pd.DataFrame) -> pd.DataFrame:
    """汇总后重算总海外/总库存与周转天数（避免对比率求和）。"""
    out = grouped
    if {"总在途库存", "总可售库存"}.issubset(out.columns):
        out["总海外库存"] = out["总在途库存"] + out["总可售库存"]
    if {"总在途库存", "总可售库存", "总计划库存"}.issubset(out.columns):
        out["总库存"] = (
            out["总在途库存"] + out["总可售库存"] + out["总计划库存"]
        )

    # 销量为 0 时结果为 NaN
    if "参考月销量" in out.columns:
        sales_safe = out["参考月销量"].mask(out["参考月销量"].eq(0))
        if "总海外库存" in out.columns:
            out["海外周转-月"] = out["总海外库存"] / sales_safe
        if "总库存" in out.columns:
            out["总库存周转-月"] = out["总库存"] / sales_safe
    return out


def _build_product_turnover(
    sku_df: pd.DataFrame,
    *,
    group_keys: list[str],
    drop_cols: tuple[str, ...] = (),
) -> pd.DataFrame:
    """
    由各销售平台SKU库存动销明细汇总商品级周转明细：
    商品ID 为空 → SKU；按 group_keys 分组，数值列求和（货值除外），周转天数重算。
    """
    drop_set = frozenset(drop_cols)
    cols = [c for c in _EXPORT_COLS if c in sku_df.columns and c not in drop_set]
    if sku_df.empty:
        return pd.DataFrame(columns=cols)

    src = _fill_blank_product_id(sku_df)
    for key in group_keys:
        if key not in src.columns:
            raise KeyError(f"商品级汇总缺少分组列：{key}")

    agg: dict[str, str | Callable] = {}
    for col in _PRODUCT_FIRST_COLS:
        if col in src.columns and col not in group_keys and col not in drop_set:
            agg[col] = _first_prefer_status if col == "产品状态" else "first"

    sum_candidates = (
        list(_INT_COLS) + [c for c in _FLOAT_COLS if c not in _PRODUCT_SKIP_SUM]
    )
    for col in sum_candidates:
        if col in src.columns and col not in group_keys and col not in agg:
            agg[col] = "sum"

    grouped = src.groupby(group_keys, as_index=False, dropna=False).agg(agg)
    grouped = _recompute_inventory_metrics(grouped)
    grouped = _coerce_numeric_columns(grouped)
    grouped = _compute_value_metrics(grouped)
    out_cols = [c for c in _EXPORT_COLS if c in grouped.columns and c not in drop_set]
    return grouped[out_cols]


# ---------------------------------------------------------------------------
# 透视
# ---------------------------------------------------------------------------

def _normalize_pivot_dim(series: pd.Series, *, blank_label: str) -> pd.Series:
    """空维度值替换为 无/nobody 等标签。"""
    s = series.astype(object).where(series.notna(), "")
    s = s.map(lambda v: str(v).strip())
    return s.mask(s.isin(_NA_TOKENS), blank_label)


def _prepare_pivot_source(sku_df: pd.DataFrame) -> pd.DataFrame:
    """透视前准备：数量列转数值，并预计算行级货值金额。"""
    src = sku_df.copy()
    for col in _PIVOT_QTY_COLS:
        if col not in src.columns:
            src[col] = 0
        src[col] = _num(src, col)

    unit_cost = _num(src, "货值/个")
    if "总海外库存" in src.columns:
        overseas = _num(src, "总海外库存")
    else:
        overseas = _num(src, "总在途库存") + _num(src, "总可售库存")

    if "总库存" in src.columns:
        total_inv = _num(src, "总库存")
    else:
        total_inv = overseas + _num(src, "总计划库存")

    src["货值(在途+在库)"] = unit_cost * overseas
    src["货值(整体库存)"] = unit_cost * total_inv
    return src


def _attach_pivot_turnover(df: pd.DataFrame) -> pd.DataFrame:
    """
    海外库存周转 = (在途禁调+可售禁调+在途可调+可售可调) / 参考月销量
    整体库存周转 = (计划禁调+在途禁调+可售禁调+计划可调+在途可调+可售可调) / 参考月销量
    参考月销量为 0 → 0
    """
    out = df.copy()
    b = _num(out, "参考月销量")
    c = _num(out, "计划库存-禁调")
    d = _num(out, "在途库存-禁调")
    e = _num(out, "可售库存-禁调")
    f = _num(out, "计划库存-可调")
    g = _num(out, "在途库存-可调")
    h = _num(out, "可售库存-可调")
    b_safe = b.mask(b.eq(0))
    out["海外库存周转"] = ((d + e + g + h) / b_safe).fillna(0)
    out["整体库存周转"] = ((c + d + e + f + g + h) / b_safe).fillna(0)
    return out


def _build_one_pivot_section(
    sku_df: pd.DataFrame,
    *,
    source_col: str,
    dim_label: str,
) -> pd.DataFrame:
    """按单一维度汇总一块透视表（含总计行）。"""
    headers = [dim_label, *_PIVOT_VALUE_COLS]
    if sku_df.empty or source_col not in sku_df.columns:
        return pd.DataFrame(columns=headers)

    src = _prepare_pivot_source(sku_df)
    blank_label = _PIVOT_BLANK_LABELS.get(source_col, "空白")
    src["_dim"] = _normalize_pivot_dim(src[source_col], blank_label=blank_label)

    sum_cols = {col: "sum" for col in _PIVOT_SUM_COLS}
    grouped = src.groupby("_dim", as_index=False, dropna=False).agg(sum_cols)
    grouped = grouped.rename(columns={"_dim": dim_label})
    grouped = grouped.sort_values(by=dim_label, kind="mergesort").reset_index(drop=True)
    if dim_label == "销售平台":
        grouped = _put_platform_all_last(grouped, col=dim_label)

    total = {dim_label: "总计"}
    for col in _PIVOT_SUM_COLS:
        total[col] = _num(grouped, col).sum()
    out = pd.concat([grouped, pd.DataFrame([total])], ignore_index=True)
    out = _attach_pivot_turnover(out)
    return out[headers]


def _build_pivot_sections(sku_df: pd.DataFrame) -> list[pd.DataFrame]:
    """生成三块透视：销售平台 / 产品状态 / 销售负责人。"""
    return [
        _build_one_pivot_section(sku_df, source_col=src, dim_label=label)
        for label, src in _PIVOT_SECTIONS
    ]


# ---------------------------------------------------------------------------
# Excel 写出与 main
# ---------------------------------------------------------------------------

def _autosize_columns(ws) -> None:
    for col_cells in ws.columns:
        letter = get_column_letter(col_cells[0].column)
        max_w = 0
        for cell in col_cells:
            max_w = max(max_w, _display_width(cell.value))
        ws.column_dimensions[letter].width = min(
            _COL_WIDTH_MAX, max(_COL_WIDTH_MIN, max_w + _COL_WIDTH_PAD)
        )


def _apply_sheet_styles(ws, columns: list[str]) -> None:
    """表头样式、禁调列灰底、边框、居中、冻结首行、自动列宽。"""
    dir_idxs = {i for i, name in enumerate(columns, start=1) if name in _DIR_COLS}
    float_idxs = {i for i, name in enumerate(columns, start=1) if name in _FLOAT_COLS}
    int_idxs = {i for i, name in enumerate(columns, start=1) if name in _INT_COLS}

    ws.row_dimensions[1].height = 28
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions

    for row in ws.iter_rows(min_row=1, max_row=ws.max_row, max_col=ws.max_column):
        for cell in row:
            cell.border = _THIN_BORDER
            cell.font = _HEADER_FONT if cell.row == 1 else _BODY_FONT
            cell.alignment = _HEADER_ALIGN if cell.row == 1 else _BODY_ALIGN

            if cell.row == 1:
                cell.fill = _HEADER_FILL
            elif cell.column in dir_idxs:
                cell.fill = _DIR_FILL

            if cell.row > 1 and cell.column in float_idxs and cell.value is not None:
                cell.number_format = "0.00"
            elif cell.row > 1 and cell.column in int_idxs and cell.value is not None:
                cell.number_format = "0"

    _autosize_columns(ws)


def _write_pivot_sheet(ws, sections: list[pd.DataFrame]) -> None:
    """把三块透视纵向写入同一 sheet，块之间空一行。"""
    row_cursor = 1
    max_col = 1 + len(_PIVOT_VALUE_COLS)

    for section_idx, section_df in enumerate(sections):
        headers = list(section_df.columns)
        for col_idx, name in enumerate(headers, start=1):
            cell = ws.cell(row=row_cursor, column=col_idx, value=name)
            cell.font = _PIVOT_HEADER_FONT
            cell.alignment = _HEADER_ALIGN
            cell.border = _THIN_BORDER
            cell.fill = (
                _PIVOT_SALES_FILL if name == "参考月销量" else _PIVOT_HEADER_FILL
            )
        ws.row_dimensions[row_cursor].height = 30
        header_row = row_cursor
        row_cursor += 1

        for _, rec in section_df.iterrows():
            is_total = str(rec.iloc[0]) == "总计"
            for col_idx, name in enumerate(headers, start=1):
                val = rec[name]
                if pd.isna(val):
                    val = None
                elif name in _PIVOT_QTY_COLS and val is not None:
                    val = int(round(float(val)))
                elif name in _PIVOT_TURN_COLS or name in _PIVOT_AMOUNT_COLS:
                    if val is not None:
                        val = float(val)
                cell = ws.cell(row=row_cursor, column=col_idx, value=val)
                cell.font = _TOTAL_FONT if is_total else _BODY_FONT
                cell.alignment = _BODY_ALIGN
                cell.border = _THIN_BORDER
                if name in _PIVOT_TURN_COLS or name in _PIVOT_AMOUNT_COLS:
                    if val is not None:
                        cell.number_format = "0.00"
                elif name in _PIVOT_QTY_COLS and val is not None:
                    cell.number_format = "0"
            row_cursor += 1

        if section_idx < len(sections) - 1:
            row_cursor += 1

        if section_idx == 0:
            ws.freeze_panes = f"A{header_row + 1}"

    ws.column_dimensions["A"].width = 16
    for col_idx in range(2, max_col + 1):
        letter = get_column_letter(col_idx)
        max_w = _COL_WIDTH_MIN
        for row in ws.iter_rows(
            min_row=1, max_row=ws.max_row, min_col=col_idx, max_col=col_idx
        ):
            for cell in row:
                max_w = max(max_w, _display_width(cell.value))
        ws.column_dimensions[letter].width = min(
            _COL_WIDTH_MAX, max(12, max_w + _COL_WIDTH_PAD)
        )


def export_turnover(snapshot_date: date, date_label: str) -> Path:
    output_dir = Path(fr"{DESKTOP_ROOT}\{folder_name}{shared_date}\仓租")
    output_dir.mkdir(parents=True, exist_ok=True)
    output_path = output_dir / f"{date_label}{FILE_SUFFIX}"

    df = _fetch_market_turnover(snapshot_date)
    cols = [c for c in _EXPORT_COLS if c in df.columns]
    if not df.empty:
        df = df[cols]
        df = _put_platform_all_last(df)

    by_market_df = _put_platform_all_last(
        _build_product_turnover(
            df, group_keys=["销售平台", "商品ID"], drop_cols=("销售站点",)
        )
    )
    by_product_df = _build_product_turnover(
        df, group_keys=["商品ID"], drop_cols=("销售平台", "销售站点")
    )
    pivot_sections = _build_pivot_sections(df)

    detail_sheets = (
        (SHEET_SKU, df, cols),
        (SHEET_PRODUCT_BY_MARKET, by_market_df, list(by_market_df.columns)),
        (SHEET_PRODUCT, by_product_df, list(by_product_df.columns)),
    )
    with pd.ExcelWriter(output_path, engine="openpyxl") as writer:
        for sheet_name, sheet_df, sheet_cols in detail_sheets:
            sheet_df.to_excel(writer, sheet_name=sheet_name, index=False)
            _apply_sheet_styles(writer.sheets[sheet_name], sheet_cols)

        # 第 4 个 sheet：三块纵向透视
        ws = writer.book.create_sheet(SHEET_PIVOT)
        _write_pivot_sheet(ws, pivot_sections)

    pivot_rows = sum(max(0, len(s) - 1) for s in pivot_sections)  # 不含总计
    print(
        f"库存动销明细已导出：SKU {len(df)} 行 / "
        f"各销售平台商品 {len(by_market_df)} 行 / 商品 {len(by_product_df)} 行 / "
        f"透视维度行 {pivot_rows}，"
        f"快照日 {snapshot_date}，文件另存为：{output_path}"
    )
    return output_path


def main(argv: list[str] | None = None) -> int:
    parser = argparse.ArgumentParser(
        description="从 snapshot_market_turnover 导出 SKU/商品/透视 库存动销与周转明细"
    )
    parser.add_argument(
        "-d",
        "--date",
        dest="date",
        default=None,
        help="快照日（默认取 A0_set_date.report_date）。例：2026.7.18 / 2026-07-18",
    )
    args = parser.parse_args(argv)

    if args.date:
        snapshot_date = parse_snapshot_date(args.date)
        date_label = _format_ku_cun_date(snapshot_date)
    else:
        snapshot_date = report_date.date()
        date_label = ku_cun_date

    export_turnover(snapshot_date, date_label)
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
